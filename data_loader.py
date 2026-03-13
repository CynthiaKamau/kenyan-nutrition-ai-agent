"""Data loader for Kenya food dataset from Excel file"""
from typing import Dict, List, Any, Optional
import openpyxl
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class KenyaFoodDataLoader:
    """Load and organize food data from Excel file"""
    
    def __init__(self, excel_path: str = "kenya_food_dataset_with_aez_subcounty.xlsx"):
        """Initialize the data loader with Excel file path"""
        self.excel_path = Path(excel_path)
        self.data = None
        self.regional_foods = None
        self.nutrition_db = None
        self._load_data()
    
    def _load_data(self) -> None:
        """Load data from Excel file"""
        if not self.excel_path.exists():
            logger.warning(f"Excel file not found: {self.excel_path}. Using fallback initialization.")
            return
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb['Sheet']
            
            # Parse headers
            headers = [cell.value for cell in ws[1]]
            
            # Load all rows
            rows = []
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    row_data[header] = ws.cell(row=row_idx, column=col_idx).value
                rows.append(row_data)
            
            self.data = rows
            self._organize_regional_foods()
            self._organize_nutrition_db()
            logger.info(f"Loaded {len(rows)} food items from Excel")
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise
    
    def _organize_regional_foods(self) -> None:
        """Organize foods by region and category"""
        self.regional_foods = {}
        
        for row in self.data:
            region = row.get('Region', '').lower()
            category = row.get('Food category', '').lower()
            food = row.get('Food', '').lower()
            
            if not region or not category or not food:
                continue
            
            if region not in self.regional_foods:
                self.regional_foods[region] = {}
            
            if category not in self.regional_foods[region]:
                self.regional_foods[region][category] = []
            
            # Ensure no duplicates
            if food not in self.regional_foods[region][category]:
                self.regional_foods[region][category].append(food)
        
        logger.info(f"Organized {len(self.regional_foods)} regions")
    
    def _organize_nutrition_db(self) -> None:
        """Organize nutritional information by food item"""
        self.nutrition_db = {}
        
        for row in self.data:
            food = row.get('Food', '').lower()
            
            if not food or food in self.nutrition_db:
                continue
            
            nutrition = {
                "calories_per_100g": row.get('Calories per 100g', 0),
                "carbs": row.get('Carbs', 0),
                "protein": row.get('Protein', 0),
                "fat": row.get('Fat', 0),
                "fiber": row.get('Fiber', 0),
                "gi": row.get('GI', 50),
                "food_category": row.get('Food category', ''),
                "source_type": row.get('Source type', ''),
                "indigenous_vegetable": row.get('Indigenous vegetable', 'no')
            }
            
            self.nutrition_db[food] = nutrition
        
        logger.info(f"Organized nutrition data for {len(self.nutrition_db)} foods")
    
    def get_region_by_location(self, location: str) -> str:
        """Find region by location (county name or region name) from actual data"""
        location_lower = location.lower()
        
        if not self.data:
            return "central"
        
        # First, check if location is a region name directly
        if location_lower in self.regional_foods:
            return location_lower
        
        # Search through data for matching county
        for row in self.data:
            county = row.get('County', '').lower()
            if county == location_lower:
                region = row.get('Region', '').lower()
                if region:
                    return region
        
        # If direct match not found, try substring match
        for row in self.data:
            county = row.get('County', '').lower()
            if location_lower in county or county in location_lower:
                region = row.get('Region', '').lower()
                if region:
                    return region
        
        # If still not found, return the location as-is (might be a region name)
        # or default to central
        import logging
        if location_lower not in ["central", "coastal", "western", "eastern", "northern", "nyanza", "rift_valley"]:
            logging.getLogger(__name__).warning(f"Location '{location}' not found. Using central region as default.")
        
        return location_lower if location_lower in self.regional_foods else "central"
    
    def get_regional_foods(self, location: str) -> Dict[str, List[str]]:
        """Get foods available by location, auto-detecting region from data"""
        region = self.get_region_by_location(location)
        
        if self.regional_foods and region in self.regional_foods:
            return self.regional_foods[region]
        
        return {}
    
    def get_nutrition_info(self, food_item: str) -> Dict[str, Any]:
        """Get nutritional information for a food item"""
        food_lower = food_item.lower()
        
        if self.nutrition_db and food_lower in self.nutrition_db:
            return self.nutrition_db[food_lower]
        
        # Return default if not found
        return {
            "calories_per_100g": 0,
            "carbs": 0,
            "protein": 0,
            "fat": 0,
            "fiber": 0,
            "gi": 50
        }



# Singleton instance
_loader = None


def get_data_loader() -> KenyaFoodDataLoader:
    """Get or create the singleton data loader instance"""
    global _loader
    if _loader is None:
        _loader = KenyaFoodDataLoader()
    return _loader
